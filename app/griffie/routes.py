# jaarplanning.py
import io
import json
import re
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
from zoneinfo import ZoneInfo
from app.griffie import bp
from flask_login import current_user
from app.auth.utils import login_and_active_required, roles_required
from app.models import Motie, AdviceSession, User, Notification, DashboardLayout, Party, Tenant
from app import db, send_email
import datetime as dt
from sqlalchemy import nullslast
from sqlalchemy.orm import selectinload
from flask import abort
import pandas as pd
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for, current_app, g
from werkzeug.utils import secure_filename
from dateutil.parser import parse as dt_parse
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from app.griffie.forms import SpeakingTimeForm

REQUIRED_COLS = [
    "onderwerp",
    "omschrijving",
    "pfh code",
    "ontstaans datum",
    "oorspronkelijke planning",
    "huidige planning",
]

# Kleuren voor nieuwe planning-categorieën
COLUMN_COLOR_EXACT = "FFA7F3D0"   # kolomkleur voor exacte datums (groen)
ROW_COLOR_EXACT = "FFEAFCF4"
COLUMN_COLOR_MONTH = "FFFDE68A"   # kolomkleur voor maandlabels (geel)
ROW_COLOR_MONTH = "FFFFF9E1"
COLUMN_COLOR_QUARTER = "FF95B3D7" # kolomkleur voor kwartaal (blauw)
ROW_COLOR_QUARTER = "FFE6EDF6"

TENTH = Decimal("0.1")
DAY_MINUTES = Decimal("1440")
SECONDS_PER_MINUTE = Decimal("60")

SPEAKING_PRESETS = {
    "commissie_beheer": {
        "label": "Commissie Beheer",
        "total_minutes": 180,
        "chair_minutes": 27,
        "pause_minutes": 30,
        "college_minutes": 30,
        "speaker_slot_minutes": 5,
    },
    "commissie_bestuur": {
        "label": "Commissie Bestuur",
        "total_minutes": 180,
        "chair_minutes": 27,
        "pause_minutes": 30,
        "college_minutes": 30,
        "speaker_slot_minutes": 5,
    },
    "commissie_ontwikkeling": {
        "label": "Commissie Ontwikkeling",
        "total_minutes": 160,
        "chair_minutes": 24,
        "pause_minutes": 0,
        "college_minutes": 30,
        "speaker_slot_minutes": 5,
    },
    "commissie_samenleving": {
        "label": "Commissie Samenleving",
        "total_minutes": 160,
        "chair_minutes": 24,
        "pause_minutes": 10,
        "college_minutes": 30,
        "speaker_slot_minutes": 5,
    },
}

DAY_NAMES_NL = [
    "maandag",
    "dinsdag",
    "woensdag",
    "donderdag",
    "vrijdag",
    "zaterdag",
    "zondag",
]

MONTH_NAMES_NL = [
    "januari",
    "februari",
    "maart",
    "april",
    "mei",
    "juni",
    "juli",
    "augustus",
    "september",
    "oktober",
    "november",
    "december",
]

APPLICATION_REGISTRY = [
    {
        "slug": "spreektijden",
        "title": "Spreektijden",
        "description": "Bereken spreektijden per fractie op basis van presets of eigen instellingen en exporteer direct naar PDF.",
        "icon": "fa-microphone-lines",
        "chip": "Vergaderingen",
        "gradient": "from-sky-500 via-blue-500 to-indigo-500",
        "endpoints": ("griffie.spreektijden",),
        "default_roles": ["griffie", "bestuursadviseur"],
    },
    {
        "slug": "actielijst",
        "title": "Jaarplanning",
        "description": "Zet ruwe Excel-overzichten om naar een gestructureerde jaarplanning met kleurcodes en slimme sortering.",
        "icon": "fa-calendar-days",
        "chip": "Planning",
        "gradient": "from-emerald-500 via-teal-500 to-cyan-500",
        "endpoints": ("griffie.jaarplanning",),
        "default_roles": ["griffie", "bestuursadviseur"],
    },
]

GLOBAL_TENANT_SLUG = "__global__"


def _application_role_mapping() -> dict:
    tenant = getattr(g, "tenant", None)
    if tenant and isinstance(getattr(tenant, "settings", None), dict):
        return tenant.settings.get("application_roles") or {}
    fallback = Tenant.query.filter(Tenant.slug == GLOBAL_TENANT_SLUG).first()
    if fallback and isinstance(getattr(fallback, "settings", None), dict):
        return fallback.settings.get("application_roles") or {}
    return {}


@bp.route("/toepassingen", methods=["GET"])
@login_and_active_required
@roles_required("griffie", "bestuursadviseur")
def toepassingen():
    """Overzichtspagina met griffie-applicaties."""

    def _resolve_endpoint(*candidates: str) -> tuple[str, str | None]:
        for endpoint in candidates:
            if not endpoint:
                continue
            if endpoint in current_app.view_functions:
                try:
                    return url_for(endpoint), endpoint
                except Exception:
                    continue
        return "#", None

    role_overrides = _application_role_mapping()

    applications = []
    for app_def in APPLICATION_REGISTRY:
        url, endpoint = _resolve_endpoint(*app_def.get("endpoints", ()))
        allowed_roles = role_overrides.get(app_def["slug"], app_def.get("default_roles", [])) or []
        allowed_roles = [r.lower() for r in allowed_roles]

        is_available = endpoint is not None and url != "#"
        note = None
        if endpoint and app_def.get("note_map"):
            note = app_def["note_map"].get(endpoint)

        if is_available and current_user.has_role(*allowed_roles):
            applications.append(
                {
                    "slug": app_def["slug"],
                    "title": app_def["title"],
                    "description": app_def["description"],
                    "icon": app_def["icon"],
                    "chip": app_def["chip"],
                    "gradient": app_def["gradient"],
                    "href": url,
                    "available": True,
                    "note": note,
                }
            )

    return render_template(
        "griffie/toepassingen.html",
        applications=applications,
        available_total=len(applications),
    )

def _dec(value) -> Decimal:
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value or 0))


def minutes_to_time_str(minutes) -> str:
    dec_minutes = _dec(minutes)
    sign = "-" if dec_minutes < 0 else ""
    dec_minutes = abs(dec_minutes)
    if dec_minutes == 0:
        return "0:00:00"
    total_seconds = int((dec_minutes * SECONDS_PER_MINUTE).to_integral_value(rounding=ROUND_HALF_UP))
    hours, remainder = divmod(total_seconds, 3600)
    minute_part, seconds = divmod(remainder, 60)
    return f"{sign}{hours}:{minute_part:02d}:{seconds:02d}"


def format_full_date_nl(value: date) -> str:
    if not isinstance(value, date):
        return ""
    day_name = DAY_NAMES_NL[value.weekday()]
    month_name = MONTH_NAMES_NL[value.month - 1]
    return f"{day_name} {value.day} {month_name} {value.year}"


def calculate_speaking_distribution(
    parties,
    *,
    total_minutes: int,
    chair_minutes: int,
    pause_minutes: int,
    speakers_count: int,
    speaker_slot_minutes: int,
    college_minutes: int,
):
    total_minutes_dec = _dec(total_minutes)
    chair_minutes_dec = _dec(chair_minutes)
    pause_minutes_dec = _dec(pause_minutes)
    college_minutes_dec = _dec(college_minutes)
    speaker_slot_dec = _dec(speaker_slot_minutes)
    speakers_count_dec = _dec(speakers_count)

    speaker_minutes_dec = speakers_count_dec * speaker_slot_dec
    allocated_minutes_dec = chair_minutes_dec + pause_minutes_dec + speaker_minutes_dec + college_minutes_dec
    remaining_minutes_dec = total_minutes_dec - allocated_minutes_dec

    parties_all = list(parties)
    party_count = len(parties_all)
    total_seats = sum(max(p.zetelaantal or 0, 0) for p in parties_all)

    if party_count:
        equal_pool_dec = remaining_minutes_dec * Decimal(5) / Decimal(6)
        seat_pool_dec = remaining_minutes_dec - equal_pool_dec
        base_share_dec = equal_pool_dec / Decimal(party_count)
    else:
        equal_pool_dec = seat_pool_dec = base_share_dec = Decimal("0")

    distribution = []
    for party in parties_all:
        seats = max(party.zetelaantal or 0, 0)
        if total_seats > 0:
            seat_share_dec = seat_pool_dec * Decimal(seats) / Decimal(total_seats)
        else:
            seat_share_dec = Decimal("0")
        total_share_dec = base_share_dec + seat_share_dec

        distribution.append(
            {
                "id": party.id,
                "name": party.naam,
                "abbreviation": party.afkorting,
                "list_number": party.lijstnummer_volgende,
                "seats": seats,
                "base_minutes": float(base_share_dec),
                "seat_minutes": float(seat_share_dec),
                "total_minutes": float(total_share_dec),
                "total_time": minutes_to_time_str(total_share_dec),
                "day_fraction": float(total_share_dec / DAY_MINUTES),
            }
        )

    college_info = {
        "minutes": float(college_minutes_dec),
        "time": minutes_to_time_str(college_minutes_dec),
        "day_fraction": float(college_minutes_dec / DAY_MINUTES),
    }

    return {
        "total_minutes": float(total_minutes_dec),
        "chair_minutes": float(chair_minutes_dec),
        "pause_minutes": float(pause_minutes_dec),
        "college_minutes": float(college_minutes_dec),
        "speakers_count": int(speakers_count),
        "speaker_slot_minutes": float(speaker_slot_dec),
        "speaker_minutes": float(speaker_minutes_dec),
        "allocated_minutes": float(allocated_minutes_dec),
        "remaining_minutes": float(remaining_minutes_dec),
        "equal_pool_minutes": float(equal_pool_dec),
        "seat_pool_minutes": float(seat_pool_dec),
        "distribution": distribution,
        "excluded_parties": [],
        "total_seats": total_seats,
        "college": college_info,
    }


def generate_speaking_pdf(result, *, committee_label: str, meeting_date: date):
    formatted_date = format_full_date_nl(meeting_date)
    title_text = f"Spreektijden {committee_label} van {formatted_date}"

    buffer = io.BytesIO()

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title_text, styles["Title"]),
        Spacer(1, 12),
    ]

    table_data = [
        ["Fractie", "Spreektijd (u:mm:ss)"],
    ]
    for row in result["distribution"]:
        label = row["name"]
        if row.get("abbreviation"):
            label = f"{label} ({row['abbreviation']})"
        table_data.append(
            [
                label,
                row["total_time"],
            ]
        )
    table_data.append(
        [
            "College",
            result["college"]["time"],
        ]
    )

    table = Table(table_data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = secure_filename(title_text.lower().replace(" ", "-")) or "spreektijden"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"{filename}.pdf",
    )

# ——— Helpers: triaalbepaling ———
# ===== Spreektijden-tool =====
@bp.route("/spreektijden", methods=["GET", "POST"])
@login_and_active_required
@roles_required("griffie")
def spreektijden():
    form = SpeakingTimeForm()

    preset_choices = sorted(
        [(key, preset["label"]) for key, preset in SPEAKING_PRESETS.items()],
        key=lambda item: item[1],
    )
    custom_choice = ("custom", "Speciaal evenement")
    form.committee.choices = preset_choices + [custom_choice]

    parties = (
        Party.query.filter(Party.actief.is_(True))
        .order_by(
            nullslast(Party.lijstnummer_volgende.asc()),
            Party.naam.asc(),
        )
        .all()
    )

    default_key = preset_choices[0][0] if preset_choices else custom_choice[0]
    incoming_choice = request.values.get("committee") or form.committee.data or default_key

    if request.method == "GET":
        selected_key = incoming_choice if incoming_choice in SPEAKING_PRESETS else default_key
        form.committee.data = selected_key
        preset = SPEAKING_PRESETS.get(selected_key)
        if preset:
            form.total_minutes.data = preset["total_minutes"]
            form.chair_minutes.data = preset["chair_minutes"]
            form.pause_minutes.data = preset["pause_minutes"]
            form.college_minutes.data = preset["college_minutes"]
            form.speaker_slot_minutes.data = preset.get("speaker_slot_minutes", 5)

    selected_key = form.committee.data or default_key
    committee_label_lookup = dict(preset_choices + [custom_choice])
    committee_label = committee_label_lookup.get(selected_key, "Speciaal evenement")

    preset_payload = {
        key: {
            "label": preset["label"],
            "total_minutes": preset["total_minutes"],
            "chair_minutes": preset["chair_minutes"],
            "pause_minutes": preset["pause_minutes"],
            "college_minutes": preset["college_minutes"],
            "speaker_slot_minutes": preset.get("speaker_slot_minutes", 5),
        }
        for key, preset in SPEAKING_PRESETS.items()
    }
    presets_json = json.dumps(preset_payload)

    result = None
    if form.validate_on_submit():
        total_minutes = form.total_minutes.data or 0
        chair_minutes = form.chair_minutes.data or 0
        pause_minutes = form.pause_minutes.data or 0
        college_minutes = form.college_minutes.data or 0
        speakers_count = form.speakers_count.data or 0
        speaker_slot_minutes = form.speaker_slot_minutes.data or 0

        result = calculate_speaking_distribution(
            parties,
            total_minutes=total_minutes,
            chair_minutes=chair_minutes,
            pause_minutes=pause_minutes,
            speakers_count=speakers_count,
            speaker_slot_minutes=speaker_slot_minutes,
            college_minutes=college_minutes,
        )

        if form.export_pdf.data:
            if result["remaining_minutes"] < 0:
                flash(
                    "De opgegeven tijden leveren een negatieve fractietijd op. Pas de waardes aan voordat je exporteert.",
                    "error",
                )
            elif not result["distribution"]:
                flash("Er zijn geen fracties met zetels gevonden om te verdelen.", "error")
            else:
                meeting_date = form.meeting_date.data or date.today()
                return generate_speaking_pdf(
                    result,
                    committee_label=committee_label,
                    meeting_date=meeting_date,
                )
        elif result["remaining_minutes"] < 0:
            flash(
                "Let op: de opgegeven tijden leveren een negatieve fractietijd op.",
                "warning",
            )

    return render_template(
        "griffie/spreektijden.html",
        form=form,
        result=result,
        presets_json=presets_json,
        committee_label=committee_label,
        parties=parties,
        selected_key=selected_key,
        title="Spreektijdenberekening",
    )

def normalize_planning_date(val) -> date | None:
    """
    Zet waarden uit 'Huidige planning' om naar een datum (date).
    Ondersteunt diverse representaties:
      - echte datumwaarden
      - NL maandnamen (bijv. 'maart 2026')
      - triaalachtige notaties (T1/T2/T3)
      - 'YYYY-MM', 'MM-YYYY'
    Retourneert None als het niet lukt.
    """
    if pd.isna(val):
        return None

    if isinstance(val, (datetime, pd.Timestamp, date)):
        dt = pd.to_datetime(val, errors="coerce")
        if pd.notna(dt):
            return dt.date()

    s = str(val).strip()
    if not s:
        return None

    # Probeer directe datumparse (inclusief dag)
    try:
        parsed = dt_parse(s, dayfirst=True, default=datetime(1900, 1, 1))
        if parsed.year != 1900:
            return parsed.date()
    except Exception:
        pass

    # Triaalnotatie -> eerste maand van kwartaal
    m = re.match(r"^[Tt]\s?([123])\s+(\d{4})$", s)
    if m:
        quarter = int(m.group(1))
        year = int(m.group(2))
        month = {1: 1, 2: 5, 3: 9}[quarter]
        return date(year, month, 1)

    # NL maandnamen
    nl_months = {
        "jan": 1, "januari": 1,
        "feb": 2, "februari": 2,
        "mrt": 3, "maart": 3,
        "apr": 4, "april": 4,
        "mei": 5,
        "jun": 6, "juni": 6,
        "jul": 7, "juli": 7,
        "aug": 8, "augustus": 8,
        "sep": 9, "sept": 9, "september": 9,
        "okt": 10, "oktober": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    parts = re.findall(r"[A-Za-z]+|\d{4}", s.lower())
    year = None
    month = None
    for p in parts:
        if p.isdigit() and len(p) == 4:
            year = int(p)
        elif p in nl_months:
            month = nl_months[p]
    if year and month:
        return date(year, month, 1)

    # 'YYYY-MM' / 'MM-YYYY'
    m2 = re.match(r"^(\d{4})[-/](\d{1,2})$", s)
    if m2:
        return date(int(m2.group(1)), int(m2.group(2)), 1)
    m3 = re.match(r"^(\d{1,2})[-/](\d{4})$", s)
    if m3:
        return date(int(m3.group(2)), int(m3.group(1)), 1)

    return None

def categorize_planning(target: date | None, reference: date):
    """
    Bepaalt weergave, sorteer-datum en categorie (exact/month/quarter) voor de nieuwe kolom.
    Retourneert (display_value, sort_date, category)
    """
    if target is None:
        return "", datetime(9999, 12, 31), None

    two_month_cutoff = reference + relativedelta(months=2)
    five_month_cutoff = reference + relativedelta(months=5)

    if target <= two_month_cutoff:
        month_name = MONTH_NAMES_NL[target.month - 1]
        display = f"{target.day} {month_name} {target.year}"
        sort_date = datetime.combine(target, datetime.min.time())
        return display, sort_date, "exact"

    if target <= five_month_cutoff:
        month_name = MONTH_NAMES_NL[target.month - 1]
        display = f"{month_name} {target.year}"
        sort_date = datetime.combine(target, datetime.min.time())
        return display, sort_date, "month"

    quarter = ((target.month - 1) // 3) + 1
    quarter_start_month = (quarter - 1) * 3 + 1
    display = f"Q{quarter} {target.year}"
    sort_date = datetime.combine(target, datetime.min.time())
    return display, sort_date, "quarter"

# ——— Route: upload & resultaat ———
@bp.route("/jaarplanning", methods=["GET", "POST"])
def jaarplanning():
    if request.method == "GET":
        return render_template("griffie/jaarplanning.html")

    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Kies een Excel-bestand (.xlsx).", "error")
        return redirect(url_for("griffie.jaarplanning"))

    filename = secure_filename(file.filename)
    try:
        # Lees Excel
        df_raw = pd.read_excel(file, dtype=str)  # lees alles als string; we normaliseren zelf
        # Normaliseer kolomnamen
        col_map = {c: c.strip() for c in df_raw.columns}
        df_raw.rename(columns=col_map, inplace=True)

        # Controleer verplichte kolommen (case-insensitive)
        lower_map = {c.lower(): c for c in df_raw.columns}
        missing = [c for c in REQUIRED_COLS if c.lower() not in lower_map]
        if missing:
            flash(f"Ontbrekende kolommen: {', '.join(missing)}", "error")
            return redirect(url_for("griffie.jaarplanning"))

        # Stel volgorde samen met optionele kolommen
        ordered = [
            "onderwerp",
            "omschrijving",
        ]
        if "soort behandeling" in lower_map:
            ordered.append("soort behandeling")
        ordered.extend([
            "pfh code",
            "ontstaans datum",
            "oorspronkelijke planning",
            "huidige planning",
        ])
        if "zaaknummer" in lower_map:
            ordered.append("zaaknummer")

        use_cols = [lower_map[key] for key in ordered]
        df = df_raw[use_cols].copy()

        # Nieuwe planningskolom op basis van afstand tot vandaag
        huidige_kolomnaam = lower_map["huidige planning"]
        planning_kolomnaam = "Planning overzicht"
        today_ams = datetime.now(ZoneInfo("Europe/Amsterdam")).date()

        display_values: list[str] = []
        sort_keys: list[datetime] = []
        style_categories: list[str | None] = []
        normalized_dates: list[date | None] = []
        excel_dates: list[datetime | None] = []

        for val in df[huidige_kolomnaam].tolist():
            normalized = normalize_planning_date(val)
            display, sort_dt, category = categorize_planning(normalized, today_ams)
            display_values.append(display)
            sort_keys.append(sort_dt)
            style_categories.append(category)
            normalized_dates.append(normalized)
            if normalized and normalized.day != 1:
                excel_dates.append(datetime.combine(normalized, datetime.min.time()))
            else:
                excel_dates.append(None)

        insert_position = df.columns.get_loc(huidige_kolomnaam) + 1
        df.insert(insert_position, planning_kolomnaam, display_values)
        df["_style"] = style_categories
        df["_normalized"] = normalized_dates
        df["_excel_date"] = excel_dates
        sort_series = pd.Series(sort_keys, index=df.index)
        df.sort_values(
            by=huidige_kolomnaam,
            key=lambda col: sort_series.loc[col.index],
            inplace=True,
        )
        row_styles = df["_style"].tolist()
        excel_dates_sorted = df["_excel_date"].tolist()
        df.drop(columns=["_style", "_normalized", "_excel_date"], inplace=True)

        # Schrijf naar Excel met kleuren en verborgen kolommen
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Jaarplanning")
            ws = writer.book["Jaarplanning"]
            ws.freeze_panes = "A2"

            header_font = Font(bold=True, color="FFFFFFFF")
            header_fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_border = Border(bottom=Side(style="medium", color="FFCBD5F5"))
            ws.row_dimensions[1].height = 28
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

            # Stel kolombreedtes en tekstterugloop in
            for idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(idx)
                is_primary_col = idx in (1, 2)
                ws.column_dimensions[col_letter].width = 60 if is_primary_col else 25

                if is_primary_col:
                    for cell in ws[col_letter]:
                        existing = cell.alignment or Alignment()
                        cell.alignment = Alignment(
                            horizontal=existing.horizontal,
                            vertical=existing.vertical,
                            text_rotation=existing.text_rotation,
                            wrap_text=True,
                            shrink_to_fit=existing.shrink_to_fit,
                            indent=existing.indent,
                        )

            # Zoek kolomindexen voor planningsoverzicht en zaaknummer
            header_row = 1
            planning_col_idx = None
            current_col_idx = None
            zaaknummer_col_idx = None
            for idx, cell in enumerate(ws[header_row], start=1):
                cell_value = str(cell.value).strip().lower()
                if cell_value == planning_kolomnaam.lower():
                    planning_col_idx = idx
                elif cell_value == huidige_kolomnaam.lower():
                    current_col_idx = idx
                elif cell_value == "zaaknummer":
                    zaaknummer_col_idx = idx

            style_map = {
                "exact": (ROW_COLOR_EXACT, COLUMN_COLOR_EXACT),
                "month": (ROW_COLOR_MONTH, COLUMN_COLOR_MONTH),
                "quarter": (ROW_COLOR_QUARTER, COLUMN_COLOR_QUARTER),
            }

            for row_idx, style in enumerate(row_styles, start=2):
                if not style:
                    continue
                row_color, col_color = style_map.get(style, (None, None))
                if not row_color or not col_color:
                    continue
                row_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = row_fill
                if planning_col_idx:
                    col_fill = PatternFill(start_color=col_color, end_color=col_color, fill_type="solid")
                    ws.cell(row=row_idx, column=planning_col_idx).fill = col_fill

            if zaaknummer_col_idx:
                ws.column_dimensions[get_column_letter(zaaknummer_col_idx)].hidden = True

            if current_col_idx:
                for row_idx, excel_dt in enumerate(excel_dates_sorted, start=2):
                    if not excel_dt:
                        continue
                    cell = ws.cell(row=row_idx, column=current_col_idx)
                    cell.value = excel_dt
                    cell.number_format = "dd-mm-yyyy"

            # Centreer planningsoverzicht en voeg borders toe voor leesbaarheid
            if planning_col_idx:
                planning_letter = get_column_letter(planning_col_idx)
                for cell in ws[planning_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_side = Side(style="thin", color="FFE2E8F0")
            data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).border = data_border

            # Voeg legenda toe voor kleurcodering
            legend = writer.book.create_sheet("Legenda")
            legend["A1"] = "Categorie"
            legend["B1"] = "Kolom"
            legend["C1"] = "Rij"
            legend_categories = [
                ("Binnen 2 maanden (exacte datum)", COLUMN_COLOR_EXACT, ROW_COLOR_EXACT),
                ("Maand 3 t/m 5 (maandnaam)", COLUMN_COLOR_MONTH, ROW_COLOR_MONTH),
                ("Vanaf maand 6 (kwartaal)", COLUMN_COLOR_QUARTER, ROW_COLOR_QUARTER),
            ]
            for idx, (label, col_color, row_color) in enumerate(legend_categories, start=2):
                legend[f"A{idx}"] = label
                legend[f"B{idx}"].fill = PatternFill(start_color=col_color, end_color=col_color, fill_type="solid")
                legend[f"C{idx}"].fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            legend["A1"].font = Font(bold=True)
            legend["B1"].font = Font(bold=True)
            legend["C1"].font = Font(bold=True)
            legend.column_dimensions["A"].width = 44
            legend.column_dimensions["B"].width = 24
            legend.column_dimensions["C"].width = 24
            legend.freeze_panes = "A2"

            data_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            ws.auto_filter.ref = data_range

        output.seek(0)
        # Net bestandsnaam met suffix
        base = filename.rsplit(".", 1)[0]
        download_name = f"{base}_jaarplanning.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )

    except Exception as e:
        # Voor de gebruiker: duidelijke foutmelding
        flash(f"Er ging iets mis bij het verwerken van het bestand: {e}", "error")
        return redirect(url_for("griffie.jaarplanning"))


# ===== Griffie advies: inbox & editor =====

@bp.route('/advies')
@login_and_active_required
@roles_required('griffie')
def advies_inbox():
    moties = (
        Motie.query
        .options(selectinload(Motie.indiener))
        .filter(Motie.status.ilike('Advies griffie'))
        .order_by(Motie.updated_at.desc())
        .all()
    )
    sessions = {s.motie_id: s for s in AdviceSession.query.filter(AdviceSession.motie_id.in_([m.id for m in moties])).order_by(AdviceSession.created_at.desc()).all()} if moties else {}
    reviewer_ids = [s.reviewer_id for s in sessions.values() if getattr(s, 'reviewer_id', None)] if sessions else []
    reviewers = {u.id: u for u in (User.query.filter(User.id.in_(reviewer_ids)).all() if reviewer_ids else [])}
    return render_template('griffie/advies_inbox.html', moties=moties, sessions=sessions, reviewers=reviewers)


@bp.route('/advies/<int:motie_id>', methods=['GET', 'POST'])
@login_and_active_required
@roles_required('griffie')
def advies_bewerken(motie_id: int):
    m = Motie.query.get_or_404(motie_id)
    if (m.status or '').lower() != 'advies griffie':
        flash('Deze motie staat niet in adviesmodus.', 'warning')
        return redirect(url_for('griffie.advies_inbox'))
    ses = (
        AdviceSession.query
        .filter(AdviceSession.motie_id == m.id)
        .order_by(AdviceSession.created_at.desc())
        .first()
    )
    if not ses:
        ses = AdviceSession(motie_id=m.id, requested_by_id=m.indiener_id, status='requested', draft={})
        db.session.add(ses)
        db.session.commit()

    # Claim? (via query param ?claim=1)
    if request.args.get('claim') == '1' and (ses.reviewer_id is None or ses.reviewer_id == current_user.id):
        ses.reviewer_id = current_user.id
        ses.status = 'in_progress'
        db.session.commit()

    # Opslaan wijzigingen in draft
    if request.method == 'POST':
        # Bouw een nieuw dict-object zodat JSON change tracking werkt
        # Neem bestaand commentaar mee als het niet uit het formulier komt
        _incoming_comment = request.form.get('advies_commentaar')
        _prev_comment = (ses.draft or {}).get('advies_commentaar') if isinstance(ses.draft, dict) else None
        new_draft = {
            'titel': (request.form.get('titel') or '').strip() or m.titel,
            'constaterende_dat': json.loads(request.form.get('constaterende_dat_json') or '[]'),
            'overwegende_dat': json.loads(request.form.get('overwegende_dat_json') or '[]'),
            'draagt_college_op': json.loads(request.form.get('draagt_json') or '[]'),
            'opdracht_formulering': request.form.get('opdracht_formulering') or m.opdracht_formulering,
            'status': m.status,
            'gemeenteraad_datum': request.form.get('gemeenteraad_datum') or m.gemeenteraad_datum,
            'agendapunt': request.form.get('agendapunt') or m.agendapunt,
            'advies_commentaar': (_incoming_comment if _incoming_comment is not None else (_prev_comment or '')),
        }
        ses.draft = new_draft
        if ses.reviewer_id is None:
            ses.reviewer_id = current_user.id
        action = request.form.get('action')
        if action == 'send':
            m.status = 'Geadviseerd'
            ses.status = 'returned'
            ses.returned_at = dt.datetime.utcnow()
            from app.moties.routes import _notify_advice_returned
            _notify_advice_returned(m, current_user)
            db.session.commit()
            flash('Advies teruggestuurd naar indiener.', 'success')
            return redirect(url_for('griffie.advies_inbox'))
        else:
            ses.status = 'in_progress'
            db.session.commit()
            flash('Adviesconcept opgeslagen.', 'success')
            return redirect(url_for('griffie.advies_bewerken', motie_id=m.id))

    # UI diffs
    from app.moties.routes import _motie_snapshot  # reuse function
    curr = _motie_snapshot(m)
    draft = ses.draft or curr
    def fmt_list(lst):
        return "\n".join([f"• {x}" for x in (lst or [])])
    items = [
        { 'label': 'Titel', 'old': curr.get('titel') or '', 'new': draft.get('titel') or '', 'key': 'titel' },
        { 'label': 'Vergaderdatum', 'old': curr.get('gemeenteraad_datum') or '', 'new': draft.get('gemeenteraad_datum') or '', 'key': 'gemeenteraad_datum' },
        { 'label': 'Agendapunt', 'old': curr.get('agendapunt') or '', 'new': draft.get('agendapunt') or '', 'key': 'agendapunt' },
        { 'label': 'Constaterende dat', 'old': fmt_list(curr.get('constaterende_dat')), 'new': fmt_list(draft.get('constaterende_dat')), 'key': 'constaterende_dat' },
        { 'label': 'Overwegende dat', 'old': fmt_list(curr.get('overwegende_dat')), 'new': fmt_list(draft.get('overwegende_dat')), 'key': 'overwegende_dat' },
        { 'label': 'Draagt het college op', 'old': fmt_list(curr.get('draagt_college_op')), 'new': fmt_list(draft.get('draagt_college_op')), 'key': 'draagt_college_op' },
        { 'label': 'Opdracht', 'old': curr.get('opdracht_formulering') or '', 'new': draft.get('opdracht_formulering') or '', 'key': 'opdracht_formulering' },
    ]
    reviewer = User.query.get(ses.reviewer_id) if ses.reviewer_id else None
    users_griffie = User.query.filter(User.actief.is_(True)).all()
    users_griffie = [u for u in users_griffie if u.has_role('griffie') or u.has_role('superadmin')]
    return render_template('griffie/advies_bewerken.html', motie=m, sessie=ses, reviewer=reviewer, items=items, users_griffie=users_griffie)


@bp.route('/advies/<int:motie_id>/assign', methods=['POST'])
@login_and_active_required
@roles_required('griffie')
def advies_toewijzen(motie_id: int):
    m = Motie.query.get_or_404(motie_id)
    ses = AdviceSession.query.filter(AdviceSession.motie_id == m.id).order_by(AdviceSession.created_at.desc()).first()
    if not ses:
        abort(404)
    uid = request.form.get('reviewer_id')
    try:
        uid = int(uid)
    except Exception:
        flash('Ongeldige toewijzing.', 'danger')
        return redirect(url_for('griffie.advies_bewerken', motie_id=m.id))
    user = User.query.get(uid)
    if not user or not user.has_role('griffie', 'superadmin'):
        flash('Alleen griffie/superadmin kan worden toegewezen.', 'danger')
        return redirect(url_for('griffie.advies_bewerken', motie_id=m.id))
    ses.reviewer_id = user.id
    if ses.status == 'requested':
        ses.status = 'in_progress'
    db.session.commit()
    flash('Toegewezen.', 'success')
    return redirect(url_for('griffie.advies_bewerken', motie_id=m.id))


@bp.route('/advies/<int:motie_id>/klaar', methods=['POST'])
@login_and_active_required
@roles_required('griffie')
def advies_klaar(motie_id: int):
    m = Motie.query.get_or_404(motie_id)
    ses = AdviceSession.query.filter(AdviceSession.motie_id == m.id).order_by(AdviceSession.created_at.desc()).first()
    if not ses:
        abort(404)
    m.status = 'Geadviseerd'
    ses.status = 'returned'
    ses.returned_at = dt.datetime.utcnow()
    # notify indiener (maak eerst notification aan, commit daarna)
    from app.moties.routes import _notify_advice_returned
    _notify_advice_returned(m, current_user)
    db.session.commit()
    flash('Advies teruggestuurd naar indiener.', 'success')
    return redirect(url_for('griffie.advies_inbox'))


# ===== Griffie indienen-pagina =====
@bp.route('/indienen')
@login_and_active_required
@roles_required('griffie')
def indienen_index():
    moties = (
        Motie.query
        .filter(Motie.status.ilike('Klaar om in te dienen'))
        .order_by(Motie.updated_at.desc())
        .all()
    )
    return render_template('griffie/indienen.html', moties=moties)


# ===== Griffie dashboard (drag & drop) =====
@bp.route('/dashboard', methods=['GET'])
@login_and_active_required
@roles_required('griffie')
def dashboard_builder():
    cfg = (
        DashboardLayout.query
        .filter(DashboardLayout.user_id == current_user.id, DashboardLayout.context == 'griffie')
        .first()
    )
    # Standaard widgets lay-out
    default_layout = {
        "widgets": [
            {"id": "to_advise", "title": "Te adviseren", "x": 0, "y": 0, "w": 6, "h": 4},
            {"id": "ready_submit", "title": "Klaar om in te dienen", "x": 6, "y": 0, "w": 6, "h": 4},
            {"id": "my_claims", "title": "Mijn claims", "x": 0, "y": 4, "w": 6, "h": 3},
            {"id": "stats", "title": "Statistiek", "x": 6, "y": 4, "w": 6, "h": 3},
        ]
    }
    layout = cfg.layout if cfg and cfg.layout else default_layout
    return render_template('griffie/dashboard_builder.html', layout=layout)


@bp.route('/dashboard/save', methods=['POST'])
@login_and_active_required
@roles_required('griffie')
def dashboard_save():
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict) or 'widgets' not in data:
        return {"ok": False, "error": "Ongeldig formaat"}, 400
    cfg = (
        DashboardLayout.query
        .filter(DashboardLayout.user_id == current_user.id, DashboardLayout.context == 'griffie')
        .first()
    )
    if not cfg:
        cfg = DashboardLayout(user_id=current_user.id, context='griffie', layout={})
        db.session.add(cfg)
    cfg.layout = data
    db.session.commit()
    return {"ok": True}


@bp.route('/dashboard/view', methods=['GET'])
@login_and_active_required
@roles_required('griffie')
def dashboard_view():
    cfg = (
        DashboardLayout.query
        .filter(DashboardLayout.user_id == current_user.id, DashboardLayout.context == 'griffie')
        .first()
    )
    default_layout = {
        "widgets": [
            {"id": "to_advise", "title": "Te adviseren", "x": 0, "y": 0, "w": 6, "h": 4},
            {"id": "ready_submit", "title": "Klaar om in te dienen", "x": 6, "y": 0, "w": 6, "h": 4},
            {"id": "my_claims", "title": "Mijn claims", "x": 0, "y": 4, "w": 6, "h": 3},
            {"id": "stats", "title": "Statistiek", "x": 6, "y": 4, "w": 6, "h": 3},
        ]
    }
    layout = cfg.layout if cfg and cfg.layout else default_layout

    # Data per widget
    to_advise = (
        Motie.query
        .options(selectinload(Motie.indiener))
        .filter(Motie.status.ilike('Advies griffie'))
        .order_by(Motie.updated_at.desc())
        .limit(10)
        .all()
    )
    ready_submit = (
        Motie.query
        .options(selectinload(Motie.indiener))
        .filter(Motie.status.ilike('Klaar om in te dienen'))
        .order_by(Motie.updated_at.desc())
        .limit(10)
        .all()
    )
    my_claims = (
        db.session.query(AdviceSession, Motie)
        .join(Motie, Motie.id == AdviceSession.motie_id)
        .options(selectinload(Motie.indiener))
        .filter(AdviceSession.reviewer_id == current_user.id)
        .order_by(AdviceSession.updated_at.desc())
        .limit(10)
        .all()
    )
    counts = {
        'to_advise': len(to_advise),
        'ready_submit': len(ready_submit),
        'my_claims': len(my_claims),
    }

    widget_data = {
        'to_advise': to_advise,
        'ready_submit': ready_submit,
        'my_claims': my_claims,
        'stats': counts,
    }

    return render_template('griffie/dashboard.html', layout=layout, widget_data=widget_data)





